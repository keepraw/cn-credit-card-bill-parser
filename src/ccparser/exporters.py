from __future__ import annotations

import os
import tempfile
from pathlib import Path

from .models import ReviewItem


BANK = "\u94f6\u884c"
CARD_LAST4 = "\u5361\u5c3e\u53f7"
TRANSACTION_DATE = "\u4ea4\u6613\u65e5"
POSTING_DATE = "\u5165\u8d26\u65e5"
DESCRIPTION = "\u8bf4\u660e"
TRANSACTION_CURRENCY = "\u4ea4\u6613\u5e01\u79cd"
TRANSACTION_AMOUNT = "\u4ea4\u6613\u91d1\u989d"
SETTLEMENT_CURRENCY = "\u7ed3\u7b97\u5e01\u79cd"
SETTLEMENT_AMOUNT = "\u7ed3\u7b97\u91d1\u989d"
REASON = "\u539f\u56e0"
SOURCE_FILE = "\u6765\u6e90\u6587\u4ef6"
RAW_TEXT = "\u539f\u59cb\u6587\u672c"
TRANSACTION_SHEET = "\u4ea4\u6613\u660e\u7ec6"
STANDARD_SHEET = "\u6807\u51c6\u5b57\u6bb5"
SUMMARY_SHEET = "\u5bfc\u5165\u6458\u8981"
REVIEW_SHEET = "\u590d\u6838\u4e8b\u9879"
FORMAL_TRANSACTION_COUNT = "\u6b63\u5f0f\u4ea4\u6613\u6761\u6570"
REVIEW_ITEM_COUNT = "\u590d\u6838\u4e8b\u9879\u6761\u6570"
SUMMARY_NOTE = "\u4ea4\u6613\u660e\u7ec6\u4e3a\u65e5\u5e38\u67e5\u770b\u8868\uff1b\u6807\u51c6\u5b57\u6bb5\u4fdd\u7559\u5b8c\u6574\u5ba1\u8ba1\u4fe1\u606f\uff0c\u654f\u611f\u6765\u6e90\u5217\u9ed8\u8ba4\u9690\u85cf\u3002"


UNIFIED_COLUMNS = [
    "bank",
    "card_last4",
    "statement_start",
    "statement_end",
    "transaction_date",
    "posting_date",
    "description",
    "transaction_currency",
    "transaction_amount",
    "settlement_currency",
    "settlement_amount",
    "source_file_name",
    "source_file_hash",
    "confidence",
    "raw_text",
]


FRIENDLY_COLUMNS = [
    BANK,
    CARD_LAST4,
    TRANSACTION_DATE,
    POSTING_DATE,
    DESCRIPTION,
    TRANSACTION_CURRENCY,
    TRANSACTION_AMOUNT,
    SETTLEMENT_CURRENCY,
    SETTLEMENT_AMOUNT,
]


REVIEW_COLUMNS = [
    "reason",
    "source_file_name",
    "source_file_hash",
    "bank",
    "card_last4",
    "transaction_id",
    "statement_key",
    "detail",
    "raw_text",
]


STANDARD_HIDDEN_COLUMNS = {
    "statement_start",
    "statement_end",
    "source_file_name",
    "source_file_hash",
    "raw_text",
}


REVIEW_HIDDEN_COLUMNS = {"source_file_hash", "transaction_id", "statement_key", "raw_text"}
MAX_EXCEL_TEXT_LENGTH = 1000


def export_outputs(output_dir: Path, transactions: list[dict[str, object]], review_items: list[ReviewItem]) -> None:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas and openpyxl are required to export Excel files. Run: pip install -r requirements.txt") from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    unified_frame = pd.DataFrame(transactions)
    if unified_frame.empty:
        unified_frame = pd.DataFrame(columns=UNIFIED_COLUMNS)
    else:
        unified_frame = unified_frame.reindex(columns=UNIFIED_COLUMNS)
        unified_frame = _coerce_output_types(unified_frame)
        unified_frame = _trim_excel_text(unified_frame, ["raw_text"])
        unified_frame = unified_frame.sort_values(
            by=["transaction_date", "posting_date", "bank", "card_last4", "description"],
            kind="stable",
            na_position="last",
        )
    friendly_frame = _friendly_transactions(unified_frame)

    review_frame = pd.DataFrame([item.as_output_row() for item in review_items])
    if review_frame.empty:
        review_frame = pd.DataFrame(columns=REVIEW_COLUMNS)
    else:
        review_frame = review_frame.reindex(columns=REVIEW_COLUMNS)
        review_frame = _trim_excel_text(review_frame, ["raw_text"])

    summary_frame = pd.DataFrame([
        {"item": FORMAL_TRANSACTION_COUNT, "value": len(unified_frame)},
        {"item": REVIEW_ITEM_COUNT, "value": len(review_frame)},
        {"item": DESCRIPTION, "value": SUMMARY_NOTE},
    ])

    with tempfile.TemporaryDirectory(prefix=".staging-", dir=output_dir) as staging_name:
        staging_dir = Path(staging_name)
        unified_path = staging_dir / "unified_transactions.xlsx"
        review_path = staging_dir / "review.xlsx"

        with pd.ExcelWriter(unified_path, engine="openpyxl") as writer:
            friendly_frame.to_excel(writer, sheet_name=TRANSACTION_SHEET, index=False)
            unified_frame.to_excel(writer, sheet_name=STANDARD_SHEET, index=False)
            summary_frame.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
            _format_workbook(writer.book)
            _hide_columns_by_header(writer.book[STANDARD_SHEET], STANDARD_HIDDEN_COLUMNS)

        with pd.ExcelWriter(review_path, engine="openpyxl") as writer:
            _friendly_review(review_frame).to_excel(writer, sheet_name=REVIEW_SHEET, index=False)
            review_frame.to_excel(writer, sheet_name=STANDARD_SHEET, index=False)
            _format_workbook(writer.book)
            _hide_columns_by_header(writer.book[STANDARD_SHEET], REVIEW_HIDDEN_COLUMNS)

        _validate_workbook(unified_path, {
            TRANSACTION_SHEET: FRIENDLY_COLUMNS,
            STANDARD_SHEET: UNIFIED_COLUMNS,
            SUMMARY_SHEET: ["item", "value"],
        })
        _validate_workbook(review_path, {
            REVIEW_SHEET: [REASON, SOURCE_FILE, BANK, CARD_LAST4, DESCRIPTION, RAW_TEXT],
            STANDARD_SHEET: REVIEW_COLUMNS,
        })

        os.replace(unified_path, output_dir / "unified_transactions.xlsx")
        os.replace(review_path, output_dir / "review.xlsx")


def _friendly_transactions(frame):
    if frame.empty:
        return frame.assign(**{column: [] for column in FRIENDLY_COLUMNS})[FRIENDLY_COLUMNS]
    return frame.rename(columns={
        "bank": BANK,
        "card_last4": CARD_LAST4,
        "transaction_date": TRANSACTION_DATE,
        "posting_date": POSTING_DATE,
        "description": DESCRIPTION,
        "transaction_currency": TRANSACTION_CURRENCY,
        "transaction_amount": TRANSACTION_AMOUNT,
        "settlement_currency": SETTLEMENT_CURRENCY,
        "settlement_amount": SETTLEMENT_AMOUNT,
    })[FRIENDLY_COLUMNS]


def _coerce_output_types(frame):
    import pandas as pd

    for column in ["transaction_amount", "settlement_amount", "confidence"]:
        if column in frame.columns:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    for column in ["statement_start", "statement_end", "transaction_date", "posting_date"]:
        if column in frame.columns:
            frame[column] = pd.to_datetime(frame[column], errors="coerce").dt.date
    return frame


def _trim_excel_text(frame, columns: list[str]):
    def trim(value):
        if not isinstance(value, str) or len(value) <= MAX_EXCEL_TEXT_LENGTH:
            return value
        return value[:MAX_EXCEL_TEXT_LENGTH] + "... [truncated for Excel export]"

    for column in columns:
        if column in frame.columns:
            frame[column] = frame[column].map(trim)
    return frame


def _friendly_review(frame):
    columns = [REASON, SOURCE_FILE, BANK, CARD_LAST4, DESCRIPTION, RAW_TEXT]
    if frame.empty:
        return frame.assign(**{column: [] for column in columns})[columns]
    return frame.rename(columns={
        "reason": REASON,
        "source_file_name": SOURCE_FILE,
        "bank": BANK,
        "card_last4": CARD_LAST4,
        "detail": DESCRIPTION,
        "raw_text": RAW_TEXT,
    })[columns]


def _format_workbook(workbook) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side

    header_font = Font(bold=True)
    header_border = Border(bottom=Side(style="thin", color="D9D9D9"))
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = header_font
            cell.border = header_border
            cell.alignment = header_alignment
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = body_alignment
                if isinstance(cell.value, (float, int)):
                    cell.number_format = "#,##0.00"
        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            values = [str(cell.value or "") for cell in column_cells[:80]]
            width = min(max([len(header), *(len(value) for value in values)] + [10]) + 2, 48)
            if header in {DESCRIPTION, "description", "detail", RAW_TEXT, "raw_text"}:
                width = 36
            if header in {"source_file_hash", "source_file_name"}:
                width = 24
            sheet.column_dimensions[column_cells[0].column_letter].width = width


def _hide_columns_by_header(sheet, headers: set[str]) -> None:
    for cell in sheet[1]:
        if cell.value in headers:
            sheet.column_dimensions[cell.column_letter].hidden = True


def _validate_workbook(path: Path, expected_sheets: dict[str, list[str]]) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name, expected_headers in expected_sheets.items():
            if sheet_name not in workbook.sheetnames:
                raise RuntimeError(f"Export validation failed for {path.name}: missing sheet {sheet_name!r}")
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            headers = [value for value in next(rows, ()) if value is not None]
            missing_headers = [header for header in expected_headers if header not in headers]
            if missing_headers:
                joined = ", ".join(str(header) for header in missing_headers)
                raise RuntimeError(f"Export validation failed for {path.name}/{sheet_name}: missing headers {joined}")
    finally:
        workbook.close()
