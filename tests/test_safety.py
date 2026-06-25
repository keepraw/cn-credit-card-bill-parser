from datetime import date
from decimal import Decimal

import pytest

from ccparser.db import Database
from ccparser.dedupe import assign_ids
from ccparser.exporters import export_outputs
from ccparser.models import ParsedStatement, SourceContext, Transaction


def make_statement() -> ParsedStatement:
    transaction = Transaction(
        bank="交通银行",
        card_last4="1234",
        transaction_date=date(2025, 12, 18),
        posting_date=date(2025, 12, 19),
        description="消费 星巴克",
        transaction_currency="CNY",
        transaction_amount=Decimal("35.50"),
        settlement_currency="CNY",
        settlement_amount=Decimal("35.50"),
        source_file_name="sample.html",
        source_file_hash="abc",
        confidence=0.9,
        raw_text="raw",
    )
    statement = ParsedStatement(
        bank="交通银行",
        card_last4="1234",
        statement_start=date(2025, 12, 1),
        statement_end=date(2025, 12, 31),
        transactions=[transaction],
        confidence=0.9,
        parser_name="test",
    )
    assign_ids(statement)
    return statement


def test_database_transaction_rolls_back_failed_run(tmp_path):
    db = Database(tmp_path / "statements.db")
    statement = make_statement()
    transaction = statement.transactions[0]

    with pytest.raises(RuntimeError):
        with db.transaction():
            db.upsert_statement(statement, transaction.statement_key, "abc", "sample.html", "html")
            db.insert_transaction(transaction)
            db.mark_file_processed("abc", "sample.html", "html")
            raise RuntimeError("abort")

    assert db.fetch_transactions() == []
    assert not db.file_processed("abc")
    db.close()


def test_export_validation_failure_keeps_existing_outputs(tmp_path, monkeypatch):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    unified_path = output_dir / "unified_transactions.xlsx"
    review_path = output_dir / "review.xlsx"
    unified_path.write_bytes(b"old unified")
    review_path.write_bytes(b"old review")

    def fail_validation(*args, **kwargs):
        raise RuntimeError("forced validation failure")

    monkeypatch.setattr("ccparser.exporters._validate_workbook", fail_validation)

    with pytest.raises(RuntimeError, match="forced validation failure"):
        export_outputs(output_dir, [], [])

    assert unified_path.read_bytes() == b"old unified"
    assert review_path.read_bytes() == b"old review"


def test_database_creates_query_indexes(tmp_path):
    db = Database(tmp_path / "statements.db")
    rows = db.connection.execute(
        "SELECT name FROM sqlite_master WHERE type IN ('index', 'table')"
    ).fetchall()
    names = {row["name"] for row in rows}

    assert "schema_migrations" in names
    assert "idx_statements_period" in names
    assert "idx_transactions_dates" in names
    assert "idx_transactions_bank_card_date" in names
    db.close()


def test_export_escapes_excel_formula_text(tmp_path):
    output_dir = tmp_path / "output"
    statement = make_statement()
    transaction = statement.transactions[0]
    transaction.description = '=HYPERLINK("https://example.test")'
    transaction.raw_text = '@SUM(1,1)'

    export_outputs(output_dir, [transaction.as_output_row()], [])

    from openpyxl import load_workbook

    workbook = load_workbook(output_dir / "unified_transactions.xlsx", data_only=False)
    try:
        sheet = workbook["交易明细"]
        headers = [cell.value for cell in sheet[1]]
        description_col = headers.index("说明") + 1
        assert sheet.cell(row=2, column=description_col).value.startswith("'=")

        standard = workbook["标准字段"]
        headers = [cell.value for cell in standard[1]]
        raw_col = headers.index("raw_text") + 1
        assert standard.cell(row=2, column=raw_col).value.startswith("'@")
    finally:
        workbook.close()


def test_low_confidence_transaction_goes_to_review_not_database(tmp_path, monkeypatch):
    import ccparser.pipeline as pipeline

    inbox = tmp_path / "inbox"
    inbox.mkdir()
    source_path = inbox / "sample.txt"
    source_path.write_text("transaction candidate", encoding="utf-8")

    monkeypatch.setattr(pipeline, "INBOX_DIR", inbox)
    monkeypatch.setattr(pipeline, "OUTPUT_DIR", tmp_path / "output")
    monkeypatch.setattr(pipeline, "PROCESSED_DIR", tmp_path / "processed")
    monkeypatch.setattr(pipeline, "REVIEW_FILES_DIR", tmp_path / "review_files")
    monkeypatch.setattr(pipeline, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(pipeline, "LOG_DIR", tmp_path / "logs")
    monkeypatch.setattr(pipeline, "DB_PATH", tmp_path / "data" / "statements.db")

    low_confidence = Transaction(
        bank="测试银行",
        card_last4="1234",
        transaction_date=date(2026, 6, 20),
        posting_date=date(2026, 6, 20),
        description="疑似提示文本",
        transaction_currency="CNY",
        transaction_amount=Decimal("1000.00"),
        settlement_currency="CNY",
        settlement_amount=Decimal("1000.00"),
        source_file_name="sample.txt",
        source_file_hash="lowhash",
        confidence=0.65,
        raw_text="温馨提示 2026-06-20 1000.00",
    )
    statement = ParsedStatement(
        bank="测试银行",
        card_last4="1234",
        statement_start=date(2026, 6, 1),
        statement_end=date(2026, 6, 30),
        transactions=[low_confidence],
        confidence=0.65,
        parser_name="test",
    )

    def fake_extract(path):
        return SourceContext(
            path=path,
            file_name=path.name,
            file_hash="lowhash",
            source_type="copied_text",
            text="transaction candidate",
        )

    monkeypatch.setattr(pipeline, "extract_source", fake_extract)
    monkeypatch.setattr(pipeline, "parse_statement", lambda source: statement)

    processed_count, inserted_count, _ = pipeline.run()

    db = Database(tmp_path / "data" / "statements.db")
    try:
        assert processed_count == 1
        assert inserted_count == 0
        assert db.fetch_transactions() == []
    finally:
        db.close()
